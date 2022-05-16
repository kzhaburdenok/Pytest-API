import requests
import json
import jsonpath
import openpyxl

def test_add_student_data():
    url = "http://thetestingworldapi.com/api/studentsDetails" # API url
    file = open('E:\Automation\Pytest-API\StudentData.json') # json of the API
    json_request = json.loads(file.read()) # reading json


    # Starting loop from the 2nd row 'cause 1st row is Header, row+1 because last row in range is not count
    for i in range(2, rows+1):
        cell_first_name = sheet.cell(row=i, column=1)
        cell_middle_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_birthday = sheet.cell(row=i, column=4)

        # Reading data from each cell and puting values to json request in accordance with the key name
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_birthday.value
        # Post request
        response = requests.post(url,json_request)
        # Validating
        print('\nwe added a student with these code', response.status_code)
        print('\nwe added a student with these data', response.text)
        assert response.status_code == 201, "New student is not created"



