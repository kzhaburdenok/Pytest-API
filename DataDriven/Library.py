import json
import jsonpath
import requests
import openpyxl

class Common:
    def __init__(self, FileNamePath, SheetName):
        global workbook, sheet # global means the posibility to use variables in different methods (not only localy in one)
        workbook = openpyxl.load_workbook(FileNamePath) # Loading excel file
        sheet = workbook(SheetName) # Work sheet
    
    def fetch_row_count(self):
        rows = sheet.max_row # Qty of rows
        return rows

    def fetch_columns_count(self):
        column = sheet.max_column #qty of columns
        return column

    def fetch_key_names(self):
        key_columns = sheet.max_column
        list = []
        for i in range(1, key_columns+1):
            cell = sheet.cell(row=1,column=i)
            list.insert(cell.value)
        return list

    def update_request_with_data(self,rowNumber,jsonRequest,keyList):
        col = sheet.max_column
        for i in range (1, col+1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]]=cell.value
        return jsonRequest


